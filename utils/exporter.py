import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(ranked_candidates, file_path):
    """
    Generates a beautifully styled Excel spreadsheet from ranked candidates.
    ranked_candidates: List of evaluated candidate dictionaries.
    file_path: Output target path for the .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers definition
    headers = [
        "Rank", "Candidate ID", "Candidate Name", "Overall Score", 
        "Semantic Match", "Skill Match", "Experience Score", "Projects Score", 
        "Education Score", "Certifications Score", "Recommendation", "AI Reasoning"
    ]
    
    # Apply headers
    ws.append(headers)
    
    # Style definitions
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1A1D36", end_color="1A1D36", fill_type="solid") # Dark Indigo
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Apply header formatting
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Set header row height
    ws.row_dimensions[1].height = 28
    
    # Alternating row fills (Zebra striping)
    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_odd = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Off-white
    
    # Recommendation color fills
    reco_styles = {
        "Strong Hire": {
            "font": Font(name="Segoe UI", size=10, bold=True, color="047857"), # Dark Green
            "fill": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Light Green
        },
        "Hire": {
            "font": Font(name="Segoe UI", size=10, bold=True, color="065F46"),
            "fill": PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        },
        "Consider": {
            "font": Font(name="Segoe UI", size=10, bold=True, color="92400E"), # Dark Orange
            "fill": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Orange
        },
        "Reject": {
            "font": Font(name="Segoe UI", size=10, bold=True, color="991B1B"), # Dark Red
            "fill": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
        }
    }
    
    font_body = Font(name="Segoe UI", size=10, color="1F2937")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="1F2937")
    
    # Write candidate rows
    for idx, cand in enumerate(ranked_candidates):
        row_num = idx + 2
        rec = cand.get("recommendation", "Consider")
        
        # Prepare row data
        row_data = [
            idx + 1,                                       # Rank
            cand.get("candidate_id"),                     # Candidate ID
            cand.get("name"),                             # Name
            cand.get("score_final"),                      # Overall Score
            cand.get("scores", {}).get("semantic"),       # Semantic
            cand.get("scores", {}).get("skills"),         # Skills
            cand.get("scores", {}).get("experience"),     # Experience
            cand.get("scores", {}).get("projects"),       # Projects
            cand.get("scores", {}).get("education"),       # Education
            cand.get("scores", {}).get("certifications"),  # Certifications
            rec,                                          # Recommendation
            cand.get("explanation", {}).get("text")       # Reasoning
        ]
        
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 22
        
        # Apply formatting to the row
        row_fill = fill_odd if idx % 2 == 1 else fill_even
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = font_body
            cell.fill = row_fill
            cell.border = thin_border
            
            # Alignments
            if col_num in [1, 2, 4, 5, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            # Bold for rank and final score
            if col_num in [1, 4]:
                cell.font = font_bold
                
            # Formatting numbers
            if col_num in [4, 5, 6, 7, 8, 9, 10] and cell.value is not None:
                cell.number_format = '0.00'
                
            # Special formatting for recommendation cell
            if col_num == 11:
                styles = reco_styles.get(rec, {})
                if styles:
                    cell.font = styles["font"]
                    cell.fill = styles["fill"]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Special spacing for Reasoning
        if col[0].column == 12:
            ws.column_dimensions[col_letter].width = 65
            continue
            
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save workbook
    wb.save(file_path)
    print(f"Excel report saved successfully to {file_path}")
